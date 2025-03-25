# excel.py
import openpyxl
import os

def append_answers_to_excel(results, questions, save_path="../resultados.xlsx"):
    """
    Agrega (append) los resultados de 'results' a un Excel en 'save_path'.
    Si el archivo no existe, lo crea con los encabezados:
      id, name, IA_model, <pregunta1>, <pregunta2>, ...
    'results' es una lista de diccionarios, cada uno con:
      - "name": nombre del documento
      - "IA_model": nombre del modelo
      - "responses": dict con {pregunta: respuesta}
    """
    # 1) Comprobar si existe el archivo
    if os.path.exists(save_path):
        # Abrir el workbook existente
        workbook = openpyxl.load_workbook(save_path)
        sheet = workbook.active
    else:
        # Crear un workbook nuevo
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Resultados"
        # Escribir encabezados en la fila 1
        headers = ["id", "name", "IA_model"] + questions
        for col_idx, header in enumerate(headers, start=1):
            sheet.cell(row=1, column=col_idx, value=header)

    # 2) Hallar la última fila con datos
    last_row = sheet.max_row
    if last_row < 1:
        # Si por algún motivo no hay nada, volvemos a escribir encabezados
        headers = ["id", "name", "IA_model"] + questions
        for col_idx, header in enumerate(headers, start=1):
            sheet.cell(row=1, column=col_idx, value=header)
        last_row = 1

    # 3) Escribir los nuevos resultados
    for result in results:
        # Nueva fila = la fila siguiente a la última
        new_row = sheet.max_row + 1
        # id (puedes usar new_row-1, o un contador global, etc.)
        sheet.cell(row=new_row, column=1, value=new_row - 1)
        # name
        sheet.cell(row=new_row, column=2, value=result.get("name"))
        # IA_model
        sheet.cell(row=new_row, column=3, value=result.get("IA_model"))
        # Respuestas
        for col_idx, question in enumerate(questions, start=4):
            # Si no hay respuesta, escribe "-"
            answer = result.get("responses", {}).get(question, "-")
            sheet.cell(row=new_row, column=col_idx, value=answer)

    # 4) Guardar el Excel
    workbook.save(save_path)
    return os.path.abspath(save_path)
